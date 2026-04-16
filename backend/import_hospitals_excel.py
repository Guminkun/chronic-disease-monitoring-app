"""
从 Excel 导入全国医院数据（50599家）到 hospitals 表
用法：python import_hospitals_excel.py [excel路径]
默认路径：E:\BaiduNetdiskDownload\1429 2024年全国省市县医院名单数据\全国医院数据库（50599家）.xlsx
"""
import sys
import os

sys.path.append(os.path.dirname(os.path.abspath(__file__)))

from sqlalchemy import text
from app.database import SessionLocal, engine, Base
from app import models

EXCEL_PATH = r"E:\BaiduNetdiskDownload\1429 2024年全国省市县医院名单数据\全国医院数据库（50599家）.xlsx"
BATCH_SIZE = 500  # 每批提交数量


def migrate_schema():
    """确保新字段存在（幂等，字段已存在则跳过）"""
    new_columns = [
        ("founded_year",        "VARCHAR(20)"),
        ("president",           "VARCHAR(100)"),
        ("management_type",     "VARCHAR(50)"),
        ("is_medical_insurance","VARCHAR(20)"),
        ("bed_count",           "VARCHAR(50)"),
        ("annual_outpatient",   "VARCHAR(50)"),
        ("staff_count",         "VARCHAR(50)"),
        ("departments",         "TEXT"),
        ("email",               "VARCHAR(200)"),
        ("postcode",            "VARCHAR(20)"),
        ("introduction",        "TEXT"),
    ]
    with engine.connect() as conn:
        # PostgreSQL: 查询 information_schema 获取现有列
        result = conn.execute(text(
            "SELECT column_name FROM information_schema.columns "
            "WHERE table_name = 'hospitals'"
        ))
        existing = {row[0] for row in result}
        for col_name, col_type in new_columns:
            if col_name not in existing:
                conn.execute(text(f'ALTER TABLE hospitals ADD COLUMN {col_name} {col_type}'))
                print(f"  + 新增列: {col_name}")
            else:
                print(f"  ✓ 列已存在: {col_name}")
        conn.commit()


def clean_val(v):
    """清理单元格值"""
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def import_excel(excel_path: str):
    try:
        import openpyxl
    except ImportError:
        print("请先安装 openpyxl：pip install openpyxl")
        sys.exit(1)

    print(f"读取 Excel: {excel_path}")
    wb = openpyxl.load_workbook(excel_path, read_only=True)
    ws = wb.active

    # 读取表头
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [str(h).strip() if h else "" for h in header_row]
    print(f"列头: {headers}")

    # 列名 → 索引映射
    def idx(name):
        try:
            return headers.index(name)
        except ValueError:
            return -1

    col = {
        "seq":          idx("序号"),
        "province":     idx("省"),
        "city":         idx("市"),
        "district":     idx("区县"),
        "name":         idx("医院名称"),
        "alias":        idx("医院别名"),
        "level":        idx("医院等级"),
        "type":         idx("医院类型"),
        "founded_year": idx("建院年份"),
        "president":    idx("院长姓名"),
        "management_type":      idx("经营方式"),
        "is_medical_insurance": idx("是否医保"),
        "bed_count":    idx("床位数"),
        "annual_outpatient":    idx("年门诊量"),
        "staff_count":  idx("医护人数"),
        "departments":  idx("医院科室"),
        "phone":        idx("电话"),
        "email":        idx("邮箱"),
        "address":      idx("医院地址"),
        "postcode":     idx("邮编"),
        "introduction": idx("医院简介"),
    }

    db = SessionLocal()
    total = 0
    skipped = 0
    batch = []

    try:
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            def g(key):
                i = col.get(key, -1)
                return clean_val(row[i]) if i >= 0 and i < len(row) else None

            name = g("name")
            if not name:
                skipped += 1
                continue

            hospital = models.Hospital(
                name=name,
                alias=g("alias"),
                level=g("level"),
                type=g("type"),
                address=g("address"),
                phone=g("phone"),
                province=g("province"),
                city=g("city"),
                district=g("district"),
                website=None,
                founded_year=g("founded_year"),
                president=g("president"),
                management_type=g("management_type"),
                is_medical_insurance=g("is_medical_insurance"),
                bed_count=g("bed_count"),
                annual_outpatient=g("annual_outpatient"),
                staff_count=g("staff_count"),
                departments=g("departments"),
                email=g("email"),
                postcode=g("postcode"),
                introduction=g("introduction"),
            )
            batch.append(hospital)

            if len(batch) >= BATCH_SIZE:
                db.bulk_save_objects(batch)
                db.commit()
                total += len(batch)
                batch = []
                print(f"  已导入 {total} 条...", end="\r")

        # 最后一批
        if batch:
            db.bulk_save_objects(batch)
            db.commit()
            total += len(batch)

        wb.close()
        print(f"\n导入完成：共 {total} 条，跳过 {skipped} 条（名称为空）")

    except Exception as e:
        db.rollback()
        print(f"\n导入出错（已回滚）: {e}")
        raise
    finally:
        db.close()


if __name__ == "__main__":
    excel_path = sys.argv[1] if len(sys.argv) > 1 else EXCEL_PATH

    if not os.path.exists(excel_path):
        print(f"文件不存在: {excel_path}")
        sys.exit(1)

    print("=== Step 1: 迁移表结构 ===")
    migrate_schema()

    print("\n=== Step 2: 导入数据 ===")
    import_excel(excel_path)
